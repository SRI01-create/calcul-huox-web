"""
Phase 1 — Extraction des catalogues de sections depuis l'xlsm.
Script one-shot à exécuter une fois en local.

Usage :
    python scripts/extract_catalogues.py --xlsm <chemin_vers_fichier.xlsm>

Produit dans backend/data/ :
    catalogue_H.csv  (~743 sections : IPE, HEA, HEB, HEM, IPN, PRS H, …)
    catalogue_U.csv  (~139 sections : UPE, UPN, cornières L, PRS U, …)
    catalogue_O.csv  (~283 sections : Tca, Tre, Tci)
    catalogue_X.csv  (~6  sections  : Pca, Pre, Pci)

Notes sur les désignations :
    - Préfixe "s"  → section PRS (soudée) dans tous les catalogues
    - "L"          → cornière (catalogue U)
    - "Tca/Tre/Tci" → tube creux carré / rectangulaire / circulaire
    - "Pca/Pre/Pci" → section pleine carrée / rectangulaire / circulaire
    - Tci et Pci : colonne b = None (section circulaire, pas de largeur distincte)

Unités stockées dans le CSV (ligne d'en-tête uniquement) :
    Géométrie h, b, tw/tf/t/r/d : mm
    Aires A, Av_y, Av_z          : m²
    Inerties Iy, Iz, It          : m⁴
    Modules Wel, Wpl             : m³
    Gauchissement IW             : m⁶
    Moment sectoriel Sw / Sw_w   : m⁴
    ys, ym (U seulement)         : mm
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Erreur : openpyxl manquant. Exécuter : pip install openpyxl")


# ─── Paramétrage des catalogues ─────────────────────────────────────────────
#
# col_indices : indices 0-basés des colonnes à extraire (A=0, B=1, …)
# col_names   : noms correspondants dans le CSV produit
# data_start  : première ligne de données (1=entêtes, 2=vide, 3=unités)
#

CATALOGUES = {
    "H": {
        "sheet_name": "catalogue H",
        # Colonnes A(0) à S(18) — 19 colonnes utiles
        # Exclues : T(iy), U(iz), V(ss), W(r2), X(hi), Y–BF (synonymes + sources)
        "col_indices": list(range(0, 19)),
        "col_names": [
            "designation",                              # A  Ø
            "h", "b", "tw", "tf", "r", "d",            # B–G  mm
            "A",                                        # H   m²
            "Iy", "Wel_y", "Wpl_y",                    # I–K  m⁴/m³/m³
            "Iz", "Wel_z", "Wpl_z",                    # L–N  m⁴/m³/m³
            "It", "IW", "Sw",                           # O–Q  m⁴/m⁶/m⁴
            "Av_y", "Av_z",                             # R–S  m²
        ],
    },
    "U": {
        "sheet_name": "catalogue U",
        # Colonnes A(0)–U(20) + Y(24)=ys + Z(25)=ym
        # Exclues : V(21)=ss, W(22)=r2, X(23)=hi, AA(26)–BK = synonymes + sources
        "col_indices": list(range(0, 21)) + [24, 25],
        "col_names": [
            "designation",                              # A   Ø
            "h", "b", "tw", "tf", "r", "d",            # B–G  mm
            "A",                                        # H   m²
            "Iy", "Wel_y", "Wpl_y",                    # I–K
            "Iz", "Wel_z", "Wpl_z",                    # L–N
            "It", "IW", "Sw_w",                         # O–Q  (Sw,w = moment sectoriel d'âme)
            "Av_y", "Av_z",                             # R–S
            "iy", "iz",                                 # T–U  mm  (utiles pour Ncr,TF des U)
            "ys", "ym",                                 # Y–Z  mm  (centre de cisaillement)
        ],
    },
    "O": {
        "sheet_name": "catalogue O",
        # Colonnes A(0) à N(13) — 14 colonnes
        # Exclues : O(14)=iy, P(15)=iz, Q(16)=Wt, R–AM = illustration + sources
        # Note : Wt non requis (τt calculé depuis h, b, t, It via formule de Bredt)
        "col_indices": list(range(0, 14)),
        "col_names": [
            "designation",                              # A  Ø
            "h", "b", "t",                             # B–D  mm (b=None pour Tci)
            "A",                                        # E   m²
            "Iy", "Wel_y", "Wpl_y",                    # F–H
            "Iz", "Wel_z", "Wpl_z",                    # I–K
            "It",                                       # L   m⁴
            "Av_y", "Av_z",                             # M–N  m²
        ],
    },
    "X": {
        "sheet_name": "catalogue X",
        # Colonnes A(0) à M(12) — 13 colonnes
        # Exclues : N(13)=Illustration, Q–AE = duplicata + sources
        "col_indices": list(range(0, 13)),
        "col_names": [
            "designation",                              # A  Ø
            "h", "b",                                   # B–C  mm (b=None pour Pci)
            "A",                                        # D   m²
            "Iy", "Wel_y", "Wpl_y",                    # E–G
            "Iz", "Wel_z", "Wpl_z",                    # H–J
            "It",                                       # K   m⁴
            "Av_y", "Av_z",                             # L–M  m²
        ],
    },
}

DATA_START_ROW = 4  # lignes 1=entêtes, 2=vide, 3=unités, 4=premières données


def extract_catalogue(ws, col_indices: list, col_names: list) -> list:
    """Extrait les lignes de données utiles depuis une feuille catalogue."""
    records = []
    for row_vals in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        # Lire uniquement les colonnes utiles
        vals = [row_vals[i] if i < len(row_vals) else None for i in col_indices]

        # Ignorer les lignes sans désignation
        desig = vals[0]
        if desig is None or str(desig).strip() == "":
            continue

        record = {}
        for name, val in zip(col_names, vals):
            if name == "designation":
                record[name] = str(val).strip()
            elif val is None:
                record[name] = ""          # cellule vide → chaîne vide dans le CSV
            else:
                try:
                    record[name] = float(val)
                except (ValueError, TypeError):
                    record[name] = str(val).strip()
        records.append(record)
    return records


def write_csv(records: list, col_names: list, output_path: Path) -> None:
    """Écrit les enregistrements dans un fichier CSV UTF-8."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=col_names, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def main():
    parser = argparse.ArgumentParser(
        description="Extrait les 4 catalogues de sections depuis le fichier xlsm CALCUL HUOX."
    )
    parser.add_argument(
        "--xlsm", required=True,
        help="Chemin vers CALCUL_HUOX_EC3_-_V16_-_np.xlsm",
    )
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).resolve().parent.parent / "data"),
        help="Dossier de sortie (défaut : backend/data/)",
    )
    args = parser.parse_args()

    xlsm_path = Path(args.xlsm)
    if not xlsm_path.exists():
        sys.exit(f"Fichier introuvable : {xlsm_path}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Chargement de {xlsm_path.name} …")
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)

    for cat_key, cfg in CATALOGUES.items():
        ws = wb[cfg["sheet_name"]]
        print(f"  Extraction catalogue {cat_key} ({cfg['sheet_name']}) …", end=" ", flush=True)
        records = extract_catalogue(ws, cfg["col_indices"], cfg["col_names"])
        out_path = output_dir / f"catalogue_{cat_key}.csv"
        write_csv(records, cfg["col_names"], out_path)
        print(f"{len(records)} sections → {out_path.name}")

    wb.close()
    print(f"\n✓ Extraction terminée. Fichiers CSV dans : {output_dir}")


if __name__ == "__main__":
    main()
